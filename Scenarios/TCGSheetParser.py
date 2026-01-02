# This will parse an .xlsx file and create a json structure and save it to a file.
# Uses OpenPyXL to read the .xlsx file

# This is the source of all the scenarios and the template.
google_doc_source_url = "https://docs.google.com/spreadsheets/d/10lqPLwR30qosAy0OjrQBqxEXBDEfG7U3yyYBlrTgRxU/export?format=xlsx"
file_name = "TestDownload.xlsx"

import requests
from openpyxl import load_workbook
import argparse
import json

# I need to break down the parsing into the following sections for my own sanity:
# - Scenario Info (Name, Description, Bed Status)
# - Introduction Text (Transfer Center, After Connection to OSH Doc)
# - Dispositions (4 dispositions, each with multiple responses)
# - Patient Info (Vitals T1, Vitals T2, Labs T1, Labs T2, Imaging)
# - Key Information (List of key information to discover and key interventions to request)
# - Scenario Questions (Questions, categories, answers, key info mapped to each question, etc.)

# For nicer command line argument parsing:
parser = argparse.ArgumentParser(description="Parses scenario .xlsx file and outputs JSON structure for Transfer Tycoon.")
parser.add_argument("--scenario_name", help="The SHEET name of the scenario to parse.", required=True)
parser.add_argument("--download", help="Download the latest version of the .xlsx file from Google Docs.", action="store_true")


def download_file(url, filename):
    print("Downloadeding latest from Google Docs.")
    response = requests.get(google_doc_source_url)
    if response.status_code != 200:
        raise Exception(f"Failed to download file from {google_doc_source_url}")
    with open(file_name, 'wb') as f:
        f.write(response.content)
    
    # For each sheet in the workbook, fix U+2019 characters
    wb = load_workbook(filename=filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        convert_u2019_to_ascii_each_cell(ws)
    wb.save(filename)
    wb.close()

def print_sheet_names(filename):
    wb = load_workbook(filename=filename, read_only=True)
    print("Sheet names in the workbook:")
    for sheet in wb.sheetnames:
        print(sheet)
    wb.close()

def open_workbook_sheet(filename, sheet_name):
    wb = load_workbook(filename=filename, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
    ws = wb[sheet_name]
    print(f"Opened workbook {filename}, sheet {sheet_name}")
    return ws

def get_row_given_first_col_value(ws, search_value):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == search_value:
            return row
    return None

def get_row_given_first_col_value_partial(ws, search_value):
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None and search_value in str(cell_value):
            return row
    return None

def get_number_of_rows_in_section(ws, start_row, column_to_check=2):
    row_count = 0
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_to_check).value
        if cell_value is None or cell_value == "":
            break
        row_count += 1
    return row_count

def convert_u2019_to_ascii_each_cell(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if isinstance(cell_value, str):
                if "\u2019" in cell_value:
                    ws.cell(row=row, column=col).value = cell_value.replace("\u2019", "'")
                    print(f"Replaced U+2019 in cell ({row}, {col}, sheet: {ws.title})")

def get_scenario_metadata(ws):

    # Scenario meta data structure
    scenario_metadata = {
        "scenario_name": "",
        "scenario_description": "",
        "bed_status": "", # Must be one of: "red", "yellow", "green"
        "bed_status_text": ""
    }

    row = get_row_given_first_col_value(ws, "Scenario Name")
    if row is None:
        raise ValueError("Could not find 'Scenario Name' in the first column.")

    scenario_metadata["scenario_name"] = ws.cell(row=row, column=2).value
    scenario_metadata["scenario_description"] = ws.cell(row=row+1, column=2).value
    scenario_metadata["bed_status"] = ws.cell(row=row+2, column=2).value
    scenario_metadata["bed_status_text"] = ws.cell(row=row+2, column=3).value

    # Basic validation of bed status
    valid_bed_status = ["red", "yellow", "green"]
    if scenario_metadata["bed_status"] not in valid_bed_status:
        raise ValueError(f"Invalid bed status: {scenario_metadata['bed_status']}. Must be one of {valid_bed_status}")

    return scenario_metadata


def get_dispositions(ws):

    # Disposition structure
    # disposition_entry_example = {
    #     "disposition_name": "Admit", # Must be one of: "Admit", "Send Home", "PICU", "Consult"
    #     "is_correct": False, # Boolean "Yes" or "No"
    #     "is_emtala_violation": False, # Boolean "Yes" or "No"
    #     "learner_text": "Yes, happy to accept the patient.",
    #     "responses": [
    #         {
    #             "responder_name": "Transfer Center",
    #             "responder_text": "OK, thanks everyone.",
    #         },
    #         {
    #             "responder_name": "OSH Doctor",
    #             "responder_text": "Thanks, bye.",
    #         }
    #     ]
    # }

    # Dispositions: each entry is 6 columns wide, minimum. Addtional responses are 2 columns wide.
    # There are 4+ dispositions every time, so we can loop through them.

    # Find the dispositions row first.
    dispositions_row = get_row_given_first_col_value(ws, "Dispositions")
    if dispositions_row is None:
        raise ValueError("Could not find 'Dispositions' in the first column.")

    num_dispositions = get_number_of_rows_in_section(ws, dispositions_row, column_to_check=2)

    dispositions = []
    for row in range(dispositions_row, dispositions_row + num_dispositions):

        col = 2  # Start at column B
        disposition_name = ws.cell(row=row, column=col).value
        is_correct = ws.cell(row=row, column=col+1).value
        is_emtala_violation = ws.cell(row=row, column=col+2).value
        learner_text = ws.cell(row=row, column=col+3).value

        is_correct = True if str(is_correct).strip().lower() == "yes" else False
        is_emtala_violation = True if str(is_emtala_violation).strip().lower() == "yes" else False

        # Collect responses
        responses = []
        response_col = col + 4 # 4th column after other fields

        while True:
            responder_name = ws.cell(row=row, column=response_col).value
            responder_text = ws.cell(row=row, column=response_col+1).value
            if responder_name is None or responder_text is None:
                break
            responses.append({
                "responder_name": responder_name,
                "responder_text": responder_text,
            })
            response_col += 2

        dispositions.append({
            "disposition_name": disposition_name,
            "is_correct": is_correct,
            "is_emtala_violation": is_emtala_violation,
            "learner_text": learner_text,
            "responses": responses
        })

    # Some simple validation checks
    valid_disposition_names = ["Admit", "Send Home", "PICU", "Consult"]
    if len(dispositions) < 4 or len(dispositions) > 5:
        print(f"Warning: Expected 4-5 dispositions, found {len(dispositions)}")

    for disp in dispositions:
        if disp["disposition_name"] not in valid_disposition_names:
            print(f"Warning: Invalid disposition name: {disp['disposition_name']}. Must be one of {valid_disposition_names}")
        
        if len(disp["responses"]) == 0:
            print(f"Warning: Disposition {disp['disposition_name']} has no responses.")

    return dispositions

def get_patient_info_vitals(ws):

    vitals_t1_row = get_row_given_first_col_value(ws, "Patient Info - Vitals T1")
    if vitals_t1_row is None:
        raise ValueError("Could not find 'Patient Info - Vitals T1' in the first column.")

    vitals_t2_row = get_row_given_first_col_value(ws, "Patient Info - Vitals T2")
    if vitals_t2_row is None:
        raise ValueError("Could not find 'Patient Info - Vitals T2' in the first column.")

    # Check that number of rows is 6 for each vitals section
    num_vitals_t1 = get_number_of_rows_in_section(ws, vitals_t1_row, column_to_check=3)
    num_vitals_t2 = get_number_of_rows_in_section(ws, vitals_t2_row, column_to_check=3)
    if num_vitals_t1 != 6:
        raise ValueError(f"Expected 6 rows for Vitals T1, found {num_vitals_t1}")
    if num_vitals_t2 != 6:
        raise ValueError(f"Expected 6 rows for Vitals T2, found {num_vitals_t2}")

    # Parse Vitals T1 (6 rows, Column C is vital name, Column D is value)
    vitals_t1 = {}
    for row in range(vitals_t1_row, vitals_t1_row + 6):
        vital_name = ws.cell(row=row, column=3).value
        vital_value = ws.cell(row=row, column=4).value
        vitals_t1[vital_name] = vital_value
    
    # Parse Vitals T2 (6 rows, Column C is vital name, Column D is value)
    vitals_t2 = {}
    for row in range(vitals_t2_row, vitals_t2_row + 6):
        vital_name = ws.cell(row=row, column=3).value
        vital_value = ws.cell(row=row, column=4).value
        vitals_t2[vital_name] = vital_value

    # Check that both vitals have the same keys
    if set(vitals_t1.keys()) != set(vitals_t2.keys()):
        raise ValueError("Vitals T1 and T2 have different vital names.")
    
    return {
        "vitals_t1": vitals_t1,
        "vitals_t2": vitals_t2
    }

def get_patient_info_labs(ws):

    # Labs T1 and T2 are variable length sections, we can't do much validation here.
    labs_t1_row = get_row_given_first_col_value(ws, "Patient Info - Labs T1")
    if labs_t1_row is None:
        raise ValueError("Could not find 'Patient Info - Labs T1' in the first column.")
    labs_t2_row = get_row_given_first_col_value(ws, "Patient Info - Labs T2")
    if labs_t2_row is None:
        raise ValueError("Could not find 'Patient Info - Labs T2' in the first column.")
    
    # Count number of rows in each labs section
    num_labs_t1 = get_number_of_rows_in_section(ws, labs_t1_row, column_to_check=3)
    num_labs_t2 = get_number_of_rows_in_section(ws, labs_t2_row, column_to_check=3)

    # Parse Labs T1
    labs_t1 = {}
    for row in range(labs_t1_row, labs_t1_row + num_labs_t1):
        lab_name = ws.cell(row=row, column=3).value
        lab_value = ws.cell(row=row, column=4).value
        labs_t1[lab_name] = lab_value
    # Parse Labs T2
    labs_t2 = {}
    for row in range(labs_t2_row, labs_t2_row + num_labs_t2):
        lab_name = ws.cell(row=row, column=3).value
        lab_value = ws.cell(row=row, column=4).value
        labs_t2[lab_name] = lab_value
    
    return {
        "labs_t1": labs_t1,
        "labs_t2": labs_t2
    }

def get_patient_info_imaging(ws):

    # There may be multiple imaging entries, each with a name and description.
    # The section starts with "Patient Info - Imaging" but has variable length.

    imaging_row = get_row_given_first_col_value_partial(ws, "Patient Info - Imaging")
    if imaging_row is None:
        raise ValueError("Could not find 'Patient Info - Imaging' in the first column.")

    num_imaging = get_number_of_rows_in_section(ws, imaging_row, column_to_check=2)
    imaging = {}
    for row in range(imaging_row, imaging_row + num_imaging):
        imaging_name = ws.cell(row=row, column=2).value
        imaging_description = ws.cell(row=row, column=3).value
        row_name = ws.cell(row=row, column=1).value # Used to match questions
        imaging.update({
            row_name: {
                "imaging_name": imaging_name,
                "imaging_description": imaging_description
            }
        })
    
    return imaging

def get_key_info_(ws,  section_title):

    # Key Information and Key Interventions have the same structure.
    # Column B is the name, Column C is the score.

    key_info_row = get_row_given_first_col_value(ws, section_title)
    if key_info_row is None:
        raise ValueError(f"Could not find '{section_title}' in the first column.")
    
    num_key_info = get_number_of_rows_in_section(ws, key_info_row, column_to_check=2)
    key_information = []
    
    for row in range(key_info_row, key_info_row + num_key_info):
        key_name = ws.cell(row=row, column=2).value
        key_score = ws.cell(row=row, column=3).value
        key_information.append({
            "key_name": key_name,
            "key_score": key_score
        })

    # Validate that each entry is unique
    key_names = [k["key_name"] for k in key_information]
    if len(key_names) != len(set(key_names)):
        raise ValueError(f"Duplicate entries found in '{section_title}'.")

    # Make sure none contain commas
    for k in key_information:
        if "," in k["key_name"]:
            raise ValueError(f"Key name '{k['key_name']}' in '{section_title}' contains a comma, which is not allowed.")

    return key_information

def get_key_information(ws):
    key_information = get_key_info_(ws, "Scenario Key Information")
    return key_information

def get_key_interventions(ws):
    key_interventions = get_key_info_(ws, "Scenario Key Interventions")
    return key_interventions

def get_scenario_questions(ws, key_info, key_interventions, patient_info):

    # Starts on the next row after "Scenario Questions"
    # Columns: Question Text, Category, Key words, Key Info Revealed, Patient Info Revealed, Name of Responder, Responder Text

    questions_row = get_row_given_first_col_value(ws, "Scenario Questions")
    if questions_row is None:
        raise ValueError("Could not find 'Scenario Questions' in the first column.")
    questions_row += 1  # Move to the next row

    num_questions = get_number_of_rows_in_section(ws, questions_row, column_to_check=2)

    valid_categories = ["Transfer Center", "Present Illness/Medical History", "Exams, Labs, and Imaging", "Prior Interventions", "Recommended Interventions"]

    questions = []
    for row in range(questions_row, questions_row + num_questions):
        question_text = ws.cell(row=row, column=2).value
        category = ws.cell(row=row, column=3).value
        key_words = ws.cell(row=row, column=4).value
        key_info_revealed = ws.cell(row=row, column=5).value
        patient_info_revealed = ws.cell(row=row, column=6).value

        responder_name = ws.cell(row=row, column=7).value
        responder_text = ws.cell(row=row, column=8).value

        # Validate category
        if category not in valid_categories:
            raise ValueError(f"Invalid category '{category}' in question: {question_text}. Must be one of {valid_categories}")

        # Separate key words into a list by commas
        if key_words:
            key_words = [kw.strip() for kw in key_words.split(",")]
        else:
            ValueError("Key words cannot be empty! Question text: {question_text}")

        # Separate key info revealed into a list by commas
        if key_info_revealed:
            key_info_revealed = [ki.strip() for ki in key_info_revealed.split(",")]
        else:
            key_info_revealed = []

        # Check that each key info revealed exists in the key_info
        for ki in key_info_revealed:
            if ki not in [k["key_name"] for k in key_info] and ki not in [k["key_name"] for k in key_interventions]:
                raise ValueError(f"Key info revealed '{ki}' in question '{question_text}' does not exist in key information or key interventions.")
        # Split the key info into key information and key interventions
        key_info_revealed_info = []
        key_info_revealed_interventions = []
        for ki in key_info_revealed:
            if ki in [k["key_name"] for k in key_info]:
                key_info_revealed_info.append(ki)
            elif ki in [k["key_name"] for k in key_interventions]:
                key_info_revealed_interventions.append(ki)

        # Patient info revealed is just a cell link OR maybe just the same text?
        if patient_info_revealed is not None:
            print("Patient info revealed:", patient_info_revealed)

            # Parse the cell link to get the row name, if it looks like "=A23"
            if patient_info_revealed.startswith("="):
                patient_info_revealed = patient_info_revealed[1:]  # Remove '='
                column_letter = ''.join(filter(str.isalpha, patient_info_revealed))
                row_number = int(''.join(filter(str.isdigit, patient_info_revealed)))
                column_number = ord(column_letter.upper()) - ord('A') + 1
                patient_info_revealed = ws.cell(row=row_number, column=column_number).value
                print("Resolved patient info revealed to:", patient_info_revealed)
            # else:
                # ValueError("Patient info revealed must be a cell link starting with '='. Question text: {question_text}")

        questions.append({
            "question_text": question_text,
            "category": category,
            "key_words": key_words,
            "key_info_revealed": key_info_revealed_info,
            "key_interventions_requested": key_info_revealed_interventions,
            "patient_info_revealed": patient_info_revealed,
            "responder_name": responder_name,
            "responder_text": responder_text
        })

    return questions

############# Main Script #############
def main():
    
    # Parse command line arguments
    args = parser.parse_args()
    scenario_name = args.scenario_name
    download_latest = args.download

    # First check if we need to download the latest file.
    if download_latest:
        download_file(google_doc_source_url, file_name)

    # print_sheet_names(file_name)
    ws = open_workbook_sheet(file_name, scenario_name)

    # Force calculation of dimensions so that max_row and max_column are accurate.
    # Not really sure, but this seems to be required when opening in read_only mode.
    ws.calculate_dimension(force=True)

    # print(f"Worksheet dimension: {ws.calculate_dimension(force=True)}")

    meta_data = get_scenario_metadata(ws)
    print("Scenario Metadata:", json.dumps(meta_data, indent=4))

    dispositions = get_dispositions(ws)
    # print("Dispositions:", json.dumps(dispositions, indent=4))

    patient_info_vitals = get_patient_info_vitals(ws)
    # print("Patient Info Vitals:", json.dumps(patient_info_vitals, indent=4))

    patient_info_labs = get_patient_info_labs(ws)
    # print("Patient Info Labs:", json.dumps(patient_info_labs, indent=4))

    patient_info_imaging = get_patient_info_imaging(ws)
    # print("Patient Info Imaging:", json.dumps(patient_info_imaging, indent=4))

    # Concatenate all patient info entries into a single structure
    patient_info = {}
    for key, value in patient_info_vitals.items():
        patient_info.update({key: value})
    for key, value in patient_info_labs.items():
        patient_info.update({key: value})
    for key, value in patient_info_imaging.items():
        patient_info.update({key: value})
    # print("Patient Info:", json.dumps(patient_info, indent=4))


    key_information = get_key_information(ws)
    #print("Key Information:", json.dumps(key_information, indent=4))

    key_interventions = get_key_interventions(ws)
    #print("Key Interventions:", json.dumps(key_interventions, indent=4))

    scenario_questions = get_scenario_questions(ws, key_information, key_interventions, patient_info)
    print("Scenario Questions:", json.dumps(scenario_questions, indent=4))

    # Close the workbook
    ws.parent.close()



if __name__ == "__main__":
    main()